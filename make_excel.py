"""Генерация Excel-отчёта по доходности арбитражного бота."""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint


# ── Цветовая палитра ─────────────────────────────────────────────────────────
CLR_DARK_BG   = "1E2235"   # шапка
CLR_ACCENT    = "4F81BD"   # синий акцент
CLR_GREEN_HI  = "375623"   # тёмно-зелёный текст прибыли
CLR_GREEN_BG  = "E2EFDA"   # фон прибыльных ячеек
CLR_GOLD      = "C9A227"   # заголовки бирж
CLR_ROW_ODD   = "F2F7FF"
CLR_ROW_EVEN  = "FFFFFF"
CLR_HEADER_TXT= "FFFFFF"
CLR_SUB_HDR   = "D9E1F2"
CLR_RED       = "C00000"
CLR_AMBER     = "ED7D31"

THIN = Side(style="thin", color="BFBFBF")
MED  = Side(style="medium", color="7F7F7F")
BORDER_THIN  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MED   = Border(left=MED,  right=MED,  top=MED,  bottom=MED)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _header_cell(ws, row, col, value, bg=CLR_DARK_BG, fg=CLR_HEADER_TXT,
                 size=11, bold=True, wrap=False, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = _fill(bg)
    c.font   = _font(bold=bold, color=fg, size=size)
    c.alignment = _align(h=h, wrap=wrap)
    c.border = BORDER_THIN
    return c


def _data_cell(ws, row, col, value, bg=CLR_ROW_ODD, bold=False,
               color="000000", fmt=None, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill  = _fill(bg)
    c.font  = _font(bold=bold, color=color)
    c.alignment = _align(h=h)
    c.border = BORDER_THIN
    if fmt:
        c.number_format = fmt
    return c


def _merge_header(ws, row, c1, c2, value, bg=CLR_DARK_BG, fg=CLR_HEADER_TXT,
                  size=12, bold=True):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=value)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, color=fg, size=size)
    c.alignment = _align(wrap=False)
    c.border = BORDER_MED
    return c


def fv(monthly_rate: float, weekly_deposit: float, weeks: int) -> float:
    r = (1 + monthly_rate) ** (1 / 4.333) - 1
    return weekly_deposit * ((1 + r) ** weeks - 1) / r


# ─────────────────────────────────────────────────────────────────────────────
# ЛИСТ 1 — Главная таблица доходности
# ─────────────────────────────────────────────────────────────────────────────
def sheet_main(wb: Workbook):
    ws = wb.active
    ws.title = "Доходность за год"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    deposits = [50, 100, 200, 500]
    rates    = [0.10, 0.15, 0.20, 0.30, 0.40, 0.50, 0.60, 0.75]
    r_labels = ["10%", "15%", "20%", "30%", "40%", "50%", "60%", "75%"]

    # ── Заголовок листа ──
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1,
        value="АРБИТРАЖНЫЙ БОТ — ИТОГ ЗА 12 МЕСЯЦЕВ | $100/нед × 52 недели | неограниченное число счётов")
    c.fill = _fill(CLR_DARK_BG)
    c.font = _font(bold=True, color=CLR_HEADER_TXT, size=14)
    c.alignment = _align(wrap=False)

    # ── Строка 2: группы "Еженедельный взнос" ──
    ws.row_dimensions[2].height = 30
    _header_cell(ws, 2, 1, "%/мес", CLR_DARK_BG, size=11)
    deposit_colors = ["274E13", "1F4E79", "7B2C2C", "4A235A"]
    deposit_bgs    = ["D9EAD3", "CFE2F3", "F4CCCC", "E1D5E7"]
    dep_col_start  = [2, 4, 6, 8]  # итого и прибыль

    dep_labels = [f"${d}/нед  (вложено ${d*52:,})" for d in deposits]
    for i, (dep_lbl, col_s) in enumerate(zip(dep_labels, dep_col_start)):
        ws.merge_cells(start_row=2, start_column=col_s,
                       end_row=2,   end_column=col_s + 1)
        c = ws.cell(row=2, column=col_s, value=dep_lbl)
        c.fill = _fill(deposit_bgs[i])
        c.font = _font(bold=True, color=deposit_colors[i], size=11)
        c.alignment = _align()
        c.border = BORDER_MED

    # ── Строка 3: подзаголовки ──
    ws.row_dimensions[3].height = 22
    _header_cell(ws, 3, 1, "ROI/мес", CLR_ACCENT, size=10)
    for i, col_s in enumerate(dep_col_start):
        _header_cell(ws, 3, col_s,     "Итого $",  deposit_bgs[i],
                     fg=deposit_colors[i], size=10)
        _header_cell(ws, 3, col_s + 1, "Прибыль",  deposit_bgs[i],
                     fg=deposit_colors[i], size=10)

    # ── Данные ──
    GREEN_THRESHOLDS = [
        ("F4CCCC", "C00000"),   # красный — мало
        ("FCE5CD", "B45F06"),   # оранжевый
        ("FFF2CC", "7F6000"),   # жёлтый
        ("D9EAD3", "274E13"),   # зелёный
        ("C6EFCE", "276221"),   # ярко-зелёный
        ("A8D5A2", "1D5C1A"),   # насыщенный зелёный
        ("7EC8A4", "0D3B1F"),   # тёмно-зелёный
        ("4CAF82", "051D0F"),   # очень тёмный зелёный
    ]

    for ri, (rate, r_lbl) in enumerate(zip(rates, r_labels)):
        row = ri + 4
        ws.row_dimensions[row].height = 22
        bg_row = CLR_ROW_ODD if ri % 2 == 0 else CLR_ROW_EVEN
        bg_cell, fg_cell = GREEN_THRESHOLDS[ri]

        # ROI label
        c = ws.cell(row=row, column=1, value=r_lbl)
        c.fill   = _fill(CLR_DARK_BG)
        c.font   = _font(bold=True, color=CLR_HEADER_TXT, size=12)
        c.alignment = _align()
        c.border = BORDER_THIN

        for i, (dep, col_s) in enumerate(zip(deposits, dep_col_start)):
            v = fv(rate, dep, 52)
            p = v - dep * 52

            c_tot = ws.cell(row=row, column=col_s, value=v)
            c_tot.fill = _fill(bg_cell)
            c_tot.font = _font(bold=True, color=fg_cell, size=11)
            c_tot.alignment = _align()
            c_tot.number_format = '"$"#,##0'
            c_tot.border = BORDER_THIN

            c_pr = ws.cell(row=row, column=col_s + 1, value=p)
            c_pr.fill = _fill(bg_cell)
            c_pr.font = _font(bold=False, color=fg_cell, size=10, italic=True)
            c_pr.alignment = _align()
            c_pr.number_format = '"+$"#,##0'
            c_pr.border = BORDER_THIN

    # ── Ширины колонок ──
    _set_col_width(ws, 1, 10)
    for col_s in dep_col_start:
        _set_col_width(ws, col_s,     15)
        _set_col_width(ws, col_s + 1, 14)

    # ── Легенда ──
    legend_row = len(rates) + 5
    ws.row_dimensions[legend_row].height = 18
    ws.merge_cells(start_row=legend_row, start_column=1,
                   end_row=legend_row, end_column=9)
    c = ws.cell(row=legend_row, column=1,
        value="* Расчёт: сложный процент с еженедельными взносами. "
              "Неограниченное число счётов = нет потолка ликвидности. "
              "ROI сохраняется весь год.")
    c.font = _font(italic=True, size=9, color="595959")
    c.alignment = _align(h="left")


# ─────────────────────────────────────────────────────────────────────────────
# ЛИСТ 2 — Рост по месяцам + График
# ─────────────────────────────────────────────────────────────────────────────
def sheet_growth(wb: Workbook):
    ws = wb.create_sheet("Рост по месяцам")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    checkpoints = [
        (1,4),(2,9),(3,13),(4,17),(5,22),(6,26),
        (7,30),(8,35),(9,39),(10,43),(11,47),(12,52)
    ]
    sel_rates  = [0.15, 0.30, 0.40, 0.60]
    sel_labels = ["15%/мес (консервативный)",
                  "30%/мес (умеренный)",
                  "40%/мес (хороший)",
                  "60%/мес (оптимистичный)"]
    deposit    = 100
    col_colors = ["1F4E79", "375623", "7B2C2C", "4A235A"]
    col_bgs    = ["DEEBF7", "E2EFDA", "FCE5CD", "E1D5E7"]

    # ── Заголовок ──
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1,
        value="РОСТ КАПИТАЛА ПО МЕСЯЦАМ | $100/неделю | без потолка ликвидности")
    c.fill = _fill(CLR_DARK_BG); c.font = _font(bold=True, color="FFFFFF", size=13)
    c.alignment = _align()

    # ── Шапка ──
    ws.row_dimensions[2].height = 26
    for col, txt in [(1,"Месяц"),(2,"Вложено")]:
        _header_cell(ws, 2, col, txt, CLR_ACCENT)
    for i, lbl in enumerate(sel_labels):
        _header_cell(ws, 2, i + 3, lbl, col_bgs[i], fg=col_colors[i],
                     size=10, wrap=True)

    # ── Данные ──
    for ri, (mth, wk) in enumerate(checkpoints):
        row = ri + 3
        ws.row_dimensions[row].height = 22
        bg = CLR_ROW_ODD if ri % 2 == 0 else CLR_ROW_EVEN
        inv = wk * deposit

        _data_cell(ws, row, 1, mth,   bg, bold=True,  h="center")
        _data_cell(ws, row, 2, inv,   bg, bold=False,
                   fmt='"$"#,##0', h="center")

        for i, rate in enumerate(sel_rates):
            v = fv(rate, deposit, wk)
            is_profit = v > inv
            cell_bg = col_bgs[i] if is_profit else "FFF2CC"
            c = ws.cell(row=row, column=i + 3, value=v)
            c.fill = _fill(cell_bg)
            c.font = _font(bold=(mth == 12), color=col_colors[i], size=11)
            c.alignment = _align()
            c.number_format = '"$"#,##0'
            c.border = BORDER_THIN

    # ── Итоговая строка ──
    tot_row = len(checkpoints) + 3
    ws.row_dimensions[tot_row].height = 26
    ws.merge_cells(start_row=tot_row, start_column=1,
                   end_row=tot_row,   end_column=2)
    c = ws.cell(row=tot_row, column=1, value="ИТОГ (12 мес)")
    c.fill = _fill(CLR_DARK_BG); c.font = _font(bold=True, color="FFFFFF", size=11)
    c.alignment = _align()
    c.border = BORDER_MED

    for i, rate in enumerate(sel_rates):
        v = fv(rate, deposit, 52)
        inv_total = 52 * deposit
        profit = v - inv_total
        c = ws.cell(row=tot_row, column=i + 3,
            value=f"${v:,.0f}  (+${profit:,.0f})")
        c.fill = _fill(col_bgs[i])
        c.font = _font(bold=True, color=col_colors[i], size=11)
        c.alignment = _align()
        c.border = BORDER_MED

    # ── Ширины ──
    _set_col_width(ws, 1, 10)
    _set_col_width(ws, 2, 14)
    for i in range(4):
        _set_col_width(ws, i + 3, 26)

    # ── График (LineChart) ──
    chart = LineChart()
    chart.title   = "Рост капитала ($100/нед)"
    chart.style   = 10
    chart.y_axis.title = "Капитал, $"
    chart.x_axis.title = "Месяц"
    chart.width   = 22
    chart.height  = 14

    chart_colors = ["2E75B6", "548235", "C55A11", "7030A0"]
    for i in range(4):
        data_ref = Reference(ws, min_col=i+3, min_row=2,
                             max_row=len(checkpoints)+2)
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[i].graphicalProperties.line.solidFill = chart_colors[i]
        chart.series[i].graphicalProperties.line.width = 20000

    cats = Reference(ws, min_col=1, min_row=3,
                     max_row=len(checkpoints)+2)
    chart.set_categories(cats)
    ws.add_chart(chart, "A18")


# ─────────────────────────────────────────────────────────────────────────────
# ЛИСТ 3 — Влияние улучшений
# ─────────────────────────────────────────────────────────────────────────────
def sheet_improvements(wb: Workbook):
    ws = wb.create_sheet("Влияние улучшений")
    ws.sheet_view.showGridLines = False

    steps = [
        ("Базовая версия (Poly + Kalshi только)",    0.10, "Уже работает",         "gray"),
        ("+ Poll 60с → 10с",                         0.13, "Уже реализовано",       "green"),
        ("+ 3 параллельных сделки за итерацию",      0.17, "Уже реализовано",       "green"),
        ("+ Порог прибыли 0.5%→0.3%, ×3 страниц",   0.22, "Уже реализовано",       "green"),
        ("+ Betfair подключён",                      0.35, "Нужна регистрация",      "amber"),
        ("+ Smarkets подключён",                     0.42, "Нужна регистрация",      "amber"),
        ("+ Semantic matching (sentence-transformers)",0.52,"pip install + .env=true","amber"),
        ("+ Kelly criterion (оптимальный размер)",   0.60, "Уже реализовано",        "green"),
        ("+ Несколько счётов (без потолка)",         0.60, "Нужны доп. кошельки",    "amber"),
    ]

    deposit = 100
    weeks   = 52
    invested = deposit * weeks

    bg_status = {
        "gray":  ("F2F2F2", "595959"),
        "green": ("E2EFDA", "375623"),
        "amber": ("FFF2CC", "7F6000"),
        "red":   ("FCE5CD", "C55A11"),
    }

    # ── Заголовок ──
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1,
        value="КАК КАЖДОЕ УЛУЧШЕНИЕ ВЛИЯЕТ НА ПРИБЫЛЬ | $100/нед | 12 месяцев")
    c.fill = _fill(CLR_DARK_BG); c.font = _font(bold=True, color="FFFFFF", size=13)
    c.alignment = _align()

    # ── Шапка ──
    ws.row_dimensions[2].height = 26
    headers = ["Улучшение", "ROI/мес", "Итого за год",
               "Чистая прибыль", "Прирост vs пред.", "Статус"]
    hdr_bgs = [CLR_DARK_BG, CLR_ACCENT, "375623",
               "1F4E79",   "7B2C2C",    "4A235A"]
    for col, (txt, bg) in enumerate(zip(headers, hdr_bgs), start=1):
        _header_cell(ws, 2, col, txt, bg, size=10, wrap=False)

    # ── Данные ──
    prev_fv  = 0.0
    for ri, (name, rate, status, color) in enumerate(steps):
        row = ri + 3
        ws.row_dimensions[row].height = 24
        v = fv(rate, deposit, weeks)
        profit = v - invested
        delta  = v - prev_fv if prev_fv > 0 else 0
        bg_row = CLR_ROW_ODD if ri % 2 == 0 else CLR_ROW_EVEN
        sbg, sfg = bg_status[color]

        # Название
        c = ws.cell(row=row, column=1, value=name)
        c.fill = _fill(bg_row); c.border = BORDER_THIN
        c.font = _font(bold=(color == "green"), size=10)
        c.alignment = _align(h="left")

        # ROI
        c = ws.cell(row=row, column=2, value=f"{rate*100:.0f}%/мес")
        c.fill = _fill(bg_row); c.border = BORDER_THIN
        c.font = _font(bold=True, size=11,
                       color="375623" if rate >= 0.40 else
                             "7F6000" if rate >= 0.20 else "C00000")
        c.alignment = _align()

        # Итого
        c = ws.cell(row=row, column=3, value=v)
        c.fill = _fill(bg_row); c.border = BORDER_THIN
        c.font = _font(bold=True, color="375623", size=11)
        c.alignment = _align(); c.number_format = '"$"#,##0'

        # Прибыль
        c = ws.cell(row=row, column=4, value=profit)
        c.fill = _fill(bg_row); c.border = BORDER_THIN
        c.font = _font(bold=False, color="1F4E79", size=11)
        c.alignment = _align(); c.number_format = '"+$"#,##0'

        # Прирост
        c = ws.cell(row=row, column=5,
                    value=delta if prev_fv > 0 else "—")
        c.fill = _fill(bg_row); c.border = BORDER_THIN
        if prev_fv > 0:
            c.font = _font(bold=True, color="C00000" if delta == 0 else "375623",
                           size=11)
            c.number_format = '"+$"#,##0'
        else:
            c.font = _font(color="595959", size=11)
        c.alignment = _align()

        # Статус
        c = ws.cell(row=row, column=6, value=status)
        c.fill = _fill(sbg); c.border = BORDER_THIN
        c.font = _font(bold=(color == "green"), color=sfg, size=9)
        c.alignment = _align(wrap=True)

        prev_fv = v

    # ── Итог ──
    tot_row = len(steps) + 3
    ws.row_dimensions[tot_row].height = 28
    ws.merge_cells(start_row=tot_row, start_column=1,
                   end_row=tot_row,   end_column=1)
    c = ws.cell(row=tot_row, column=1,
        value="МАКСИМАЛЬНЫЙ ПОТЕНЦИАЛ (все улучшения)")
    c.fill = _fill(CLR_DARK_BG); c.font = _font(bold=True, color="FFFFFF", size=11)
    c.alignment = _align(h="left"); c.border = BORDER_MED

    max_v = fv(0.60, deposit, weeks)
    for col, val, fmt, clr in [
        (2, "60%/мес",           None,          "FFFFFF"),
        (3, max_v,               '"$"#,##0',    "90EE90"),
        (4, max_v - invested,    '"+$"#,##0',   "90EE90"),
        (5, max_v - fv(0.10, deposit, weeks), '"+$"#,##0', "FFD700"),
        (6, "Всё активировано",  None,          "90EE90"),
    ]:
        c = ws.cell(row=tot_row, column=col, value=val)
        c.fill = _fill(CLR_DARK_BG); c.border = BORDER_MED
        c.font = _font(bold=True, color=clr, size=11)
        c.alignment = _align()
        if fmt: c.number_format = fmt

    # ── Ширины ──
    widths = [48, 14, 15, 15, 18, 26]
    for i, w in enumerate(widths, start=1):
        _set_col_width(ws, i, w)

    # ── Гистограмма ──
    chart = BarChart()
    chart.type   = "col"
    chart.title  = "Итог за год при каждом уровне улучшений"
    chart.y_axis.title = "Капитал, $"
    chart.style  = 10
    chart.width  = 26
    chart.height = 14

    data = Reference(ws, min_col=3, min_row=2,
                     max_row=len(steps) + 2)
    cats = Reference(ws, min_col=2, min_row=3,
                     max_row=len(steps) + 2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "4F81BD"

    # Градиент цвета по значению
    bar_colors = ["C0504D","C0504D","ED7D31","ED7D31",
                  "9BBB59","9BBB59","00B050","00B050","00B050"]
    for i, col in enumerate(bar_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = col
        chart.series[0].dPt.append(pt)

    ws.add_chart(chart, "A14")


# ─────────────────────────────────────────────────────────────────────────────
# ЛИСТ 4 — Один счёт vs Несколько
# ─────────────────────────────────────────────────────────────────────────────
def sheet_multiaccounts(wb: Workbook):
    ws = wb.create_sheet("Один vs Несколько счётов")
    ws.sheet_view.showGridLines = False

    def sim_single(r_peak, r_wall, dep, weeks, wall):
        b = 0.0
        for _ in range(weeks):
            r = r_wall if b >= wall else r_peak
            b = b * (1 + r) + dep
        return b

    def sim_multi(r_peak, dep, weeks, cap):
        accounts = [0.0]
        active   = 0
        opened   = 1
        for wk in range(1, weeks + 1):
            accounts = [a * (1 + r_peak) for a in accounts]
            accounts[active] += dep
            if accounts[active] >= cap and wk < weeks:
                accounts.append(0.0)
                active = len(accounts) - 1
                opened += 1
        return sum(accounts), opened

    scenarios = [
        ("Консервативный", 0.15, 0.08),
        ("Умеренный",      0.35, 0.15),
        ("Оптимистичный",  0.60, 0.15),
    ]
    dep = 100; weeks = 52; wall = 2000; cap = 1500
    invested = dep * weeks

    # ── Заголовок ──
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1,
        value="ОДИН СЧЁТ vs НЕСКОЛЬКО СЧЁТОВ | $100/нед | 52 недели")
    c.fill = _fill(CLR_DARK_BG); c.font = _font(bold=True, color="FFFFFF", size=13)
    c.alignment = _align()

    # ── Шапка ──
    ws.row_dimensions[2].height = 26
    cols_hdr = ["Сценарий", "ROI (норма)", "ROI (потолок)",
                "Один счёт", "Прибыль", "Несколько счётов",
                "Прибыль  /  Бонус"]
    bgs_hdr  = [CLR_DARK_BG, CLR_ACCENT, CLR_ACCENT,
                "C00000",    "C00000",   "375623", "375623"]
    for col, (txt, bg) in enumerate(zip(cols_hdr, bgs_hdr), start=1):
        _header_cell(ws, 2, col, txt, bg, size=10, wrap=True)

    # ── Данные ──
    s_bgs = ["DEEBF7", "E2EFDA", "FFF2CC"]
    s_fgs = ["1F4E79", "375623", "7F6000"]
    for ri, (name, r_peak_m, r_wall_m) in enumerate(scenarios):
        row = ri + 3
        ws.row_dimensions[row].height = 26
        r_p = (1 + r_peak_m) ** (1/4.333) - 1
        r_w = (1 + r_wall_m) ** (1/4.333) - 1
        s_val  = sim_single(r_p, r_w, dep, weeks, wall)
        m_val, n_acc = sim_multi(r_p, dep, weeks, cap)
        s_pr = s_val - invested
        m_pr = m_val - invested
        bonus_pct = (m_val / s_val - 1) * 100

        bg = s_bgs[ri]; fg = s_fgs[ri]

        vals = [
            (name,                     bg, fg,      True,  None),
            (f"{r_peak_m*100:.0f}%/мес", bg, "375623",True, None),
            (f"{r_wall_m*100:.0f}%/мес", bg, "C00000",False,None),
            (s_val,                    "FCE5CD","C00000",True,'"$"#,##0'),
            (s_pr,                     "FCE5CD","C00000",False,'"+$"#,##0'),
            (m_val,                    "E2EFDA","375623",True, '"$"#,##0'),
            (f"${m_pr:,.0f}  /  +{bonus_pct:.0f}%",
                                       "E2EFDA","375623",True, None),
        ]
        for col, (val, cbg, cfg, bold, fmt) in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = _fill(cbg); c.border = BORDER_THIN
            c.font = _font(bold=bold, color=cfg, size=11)
            c.alignment = _align(h="center")
            if fmt: c.number_format = fmt

    # ── Пояснение ──
    note_row = len(scenarios) + 4
    ws.row_dimensions[note_row].height = 50
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row,   end_column=7)
    c = ws.cell(row=note_row, column=1, value=(
        "Логика: один счёт упирается в потолок ликвидности (~$2,000) и ROI падает. "
        "Несколько счётов (порог $1,500/каждый) сохраняют пиковый ROI весь год. "
        "Новый счёт открывается автоматически когда предыдущий достигает $1,500. "
        "Polymarket: новый кошелёк MetaMask без KYC. Betfair/Kalshi: 1 аккаунт на человека (KYC)."
    ))
    c.fill = _fill("FFF2CC"); c.border = BORDER_MED
    c.font = _font(italic=True, size=9, color="7F6000")
    c.alignment = _align(h="left", v="top", wrap=True)

    # ── Ширины ──
    widths = [26, 14, 15, 16, 14, 18, 22]
    for i, w in enumerate(widths, start=1):
        _set_col_width(ws, i, w)


# ─────────────────────────────────────────────────────────────────────────────
# ЛИСТ 5 — Чеклист
# ─────────────────────────────────────────────────────────────────────────────
def sheet_checklist(wb: Workbook):
    ws = wb.create_sheet("Чеклист запуска")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 30

    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value="ЧЕКЛИСТ ЗАПУСКА АРБИТРАЖНОГО БОТА")
    c.fill = _fill(CLR_DARK_BG); c.font = _font(bold=True, color="FFFFFF", size=14)
    c.alignment = _align()

    ws.row_dimensions[2].height = 22
    for col, txt in [(2,"Задача"),(3,"Статус"),(4,"Приоритет"),(5,"Как сделать")]:
        _header_cell(ws, 2, col, txt, CLR_ACCENT)

    tasks = [
        # (задача, статус, приоритет, как)
        ("Polymarket API Key",                  "✅ Готово",  "ВЫСОКИЙ",  "Уже в .env"),
        ("Polymarket Secret + Passphrase",       "⏳ Нужно",   "ВЫСОКИЙ",  "Экспорт из MetaMask"),
        ("Polymarket Private Key (ETH)",         "⏳ Нужно",   "ВЫСОКИЙ",  "MetaMask → Account Details"),
        ("Регистрация Betfair",                  "⏳ Нужно",   "ВЫСОКИЙ",  "betfair.com + KYC"),
        ("Betfair App Key",                      "⏳ Нужно",   "ВЫСОКИЙ",  "developer.betfair.com"),
        ("Вписать BETFAIR_* в .env",             "⏳ Нужно",   "ВЫСОКИЙ",  "USERNAME + PASSWORD + APP_KEY"),
        ("Регистрация Smarkets",                 "⏳ Нужно",   "СРЕДНИЙ",  "smarkets.com"),
        ("Вписать SMARKETS_API_TOKEN в .env",    "⏳ Нужно",   "СРЕДНИЙ",  "Account → API Settings"),
        ("pip install sentence-transformers",    "⏳ Нужно",   "СРЕДНИЙ",  "Терминал: pip install ..."),
        ("SEMANTIC_MATCHING_ENABLED=true",       "⏳ Нужно",   "СРЕДНИЙ",  "В файле .env"),
        ("KELLY_ENABLED=true (живой счёт)",      "⏳ Нужно",   "СРЕДНИЙ",  "В .env, после пополнения"),
        ("KELLY_BANKROLL=<сумма>",               "⏳ Нужно",   "СРЕДНИЙ",  "Обновлять при пополнении"),
        ("Создать доп. MetaMask кошельки",       "⏳ Нужно",   "СРЕДНИЙ",  "Новый кошелёк → экспорт ключа"),
        ("DRY_RUN=false",                        "⏳ Нужно",   "КРИТИЧНО", "Только после проверки!"),
        ("Запустить: python run.py",             "⏳ Нужно",   "ФИНАЛ",    "Из папки arbitrage-bot/"),
    ]

    pri_colors = {
        "КРИТИЧНО": ("C00000","FCE5CD"),
        "ВЫСОКИЙ":  ("7B2C2C","F4CCCC"),
        "СРЕДНИЙ":  ("7F6000","FFF2CC"),
        "ФИНАЛ":    ("375623","E2EFDA"),
    }
    st_colors  = {
        "✅ Готово": ("375623","E2EFDA"),
        "⏳ Нужно":  ("7F6000","FFF9C4"),
    }

    for ri, (task, status, pri, how) in enumerate(tasks):
        row = ri + 3
        ws.row_dimensions[row].height = 22
        bg = CLR_ROW_ODD if ri % 2 == 0 else CLR_ROW_EVEN

        # Номер
        c = ws.cell(row=row, column=1, value=ri + 1)
        c.fill = _fill(CLR_DARK_BG); c.font = _font(bold=True, color="AAAAAA", size=9)
        c.alignment = _align(); c.border = BORDER_THIN

        # Задача
        c = ws.cell(row=row, column=2, value=task)
        c.fill = _fill(bg); c.font = _font(size=10)
        c.alignment = _align(h="left"); c.border = BORDER_THIN

        # Статус
        sfg, sbg = st_colors.get(status, ("000000","FFFFFF"))
        c = ws.cell(row=row, column=3, value=status)
        c.fill = _fill(sbg); c.font = _font(bold=True, color=sfg, size=10)
        c.alignment = _align(); c.border = BORDER_THIN

        # Приоритет
        pfg, pbg = pri_colors.get(pri, ("000000","FFFFFF"))
        c = ws.cell(row=row, column=4, value=pri)
        c.fill = _fill(pbg); c.font = _font(bold=True, color=pfg, size=10)
        c.alignment = _align(); c.border = BORDER_THIN

        # Как
        c = ws.cell(row=row, column=5, value=how)
        c.fill = _fill(bg); c.font = _font(italic=True, size=9, color="595959")
        c.alignment = _align(h="left"); c.border = BORDER_THIN


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    sheet_main(wb)
    sheet_growth(wb)
    sheet_improvements(wb)
    sheet_multiaccounts(wb)
    sheet_checklist(wb)

    out = "/Users/mac/Desktop/arbitrage_bot_analysis.xlsx"
    wb.save(out)
    print(f"Сохранено: {out}")


if __name__ == "__main__":
    main()
